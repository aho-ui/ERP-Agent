import base64
import csv
import io
import json

from asgiref.sync import sync_to_async
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

from agent.api.auth import require_auth


def _extract_pdf(data: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(data))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _extract_docx(data: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            rows.append(",".join("" if v is None else str(v) for v in row))
    return "\n".join(rows)


def _extract_csv(data: bytes) -> str:
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return "\n".join(",".join(row) for row in reader)


def _extract_audio(data: bytes, filename: str) -> str:
    import litellm
    import tempfile, os
    ext = os.path.splitext(filename)[1] or ".webm"
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(data)
        tmp_path = tmp.name
    try:
        with open(tmp_path, "rb") as f:
            response = litellm.transcription(model="whisper-1", file=f)
        return response.text
    finally:
        os.unlink(tmp_path)


@csrf_exempt
async def upload(request):
    _, _, err = await require_auth(request)
    if err:
        return err

    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    file = request.FILES.get("file")
    if not file:
        return JsonResponse({"error": "file is required"}, status=400)

    content_type = file.content_type or ""
    filename = file.name or ""
    data = file.read()

    try:
        if content_type == "application/pdf" or filename.endswith(".pdf"):
            content = await sync_to_async(_extract_pdf)(data)
        elif (
            content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            or filename.endswith(".docx")
        ):
            content = await sync_to_async(_extract_docx)(data)
        elif (
            content_type in (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel",
            )
            or filename.endswith(".xlsx")
        ):
            content = await sync_to_async(_extract_xlsx)(data)
        elif content_type in ("text/csv", "text/plain") or filename.endswith(".csv"):
            content = _extract_csv(data)
        elif content_type.startswith("image/"):
            return JsonResponse({
                "image_data": base64.b64encode(data).decode(),
                "content_type": content_type,
                "filename": filename,
            })
        elif content_type.startswith("audio/"):
            content = await sync_to_async(_extract_audio)(data, filename)
        else:
            return JsonResponse({"error": f"Unsupported file type: {content_type}"}, status=400)
    except Exception as e:
        return JsonResponse({"error": f"Failed to process file: {e}"}, status=500)

    return JsonResponse({"content": content, "filename": filename})


@csrf_exempt
@require_POST
async def export(request):
    _, _, err = await require_auth(request)
    if err:
        return err
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)
    fmt = body.get("format", "csv")
    title = body.get("title", "export")

    if fmt == "pdf":
        from agent.utils.pdf import generate_pdf_bytes
        data = generate_pdf_bytes(body["columns"], body["rows"], title=title)
        content_type = "application/pdf"
    elif fmt == "xlsx":
        from agent.utils.xlsx_export import generate_xlsx_bytes
        data = generate_xlsx_bytes(body["columns"], body["rows"])
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        from agent.utils.csv_export import generate_csv_bytes
        data = generate_csv_bytes(body["columns"], body["rows"])
        content_type = "text/csv"

    response = HttpResponse(data, content_type=content_type)
    response["Content-Disposition"] = f'attachment; filename="{title}.{fmt}"'
    return response


@csrf_exempt
@require_POST
async def po_document(request):
    _, _, err = await require_auth(request)
    if err:
        return err
    data = json.loads(request.body)
    from agent.utils.documents.po import generate_po_pdf
    pdf_bytes = generate_po_pdf(data)
    title = data.get("po_number", "po")
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{title}.pdf"'
    return response
