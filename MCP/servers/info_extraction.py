import base64
import csv
import io
import json

import litellm
from mcp.server.fastmcp import FastMCP

MODEL = "gpt-4o-mini"

mcp = FastMCP("extraction")


def health() -> bool:
    try:
        import litellm  # noqa: F401
        return True
    except Exception:
        return False


def _parse_pdf(data: bytes) -> str:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(data))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _parse_docx(data: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _parse_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            rows.append(",".join("" if v is None else str(v) for v in row))
    return "\n".join(rows)


def _parse_csv(data: bytes) -> str:
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return "\n".join(",".join(row) for row in reader)


def _extract_with_text(text: str, prompt: str) -> str:
    response = litellm.completion(
        model=MODEL,
        messages=[
            {"role": "system", "content": prompt},
            {"role": "user", "content": text},
        ],
        response_format={"type": "json_object"},
    )
    return response.choices[0].message.content


def _extract_with_image(data: bytes, content_type: str, prompt: str) -> str:
    b64 = base64.b64encode(data).decode()
    response = litellm.completion(
        model=MODEL,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image_url", "image_url": {"url": f"data:{content_type};base64,{b64}"}},
                {"type": "text", "text": prompt},
            ],
        }],
        response_format={"type": "json_object"},
    )
    return response.choices[0].message.content


def _process(file_data: str, file_type: str, prompt: str) -> str:
    data = base64.b64decode(file_data)
    if file_type.startswith("image/"):
        return _extract_with_image(data, file_type, prompt)
    if file_type == "application/pdf":
        text = _parse_pdf(data)
    elif file_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        text = _parse_docx(data)
    elif file_type in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"):
        text = _parse_xlsx(data)
    elif file_type in ("text/csv", "text/plain"):
        text = _parse_csv(data)
    else:
        return json.dumps({"error": f"Unsupported file type: {file_type}"})
    return _extract_with_text(text, prompt)


@mcp.tool()
def extract_invoice(file_data: str, file_type: str) -> str:
    prompt = (
        "Extract structured invoice data. "
        "Return JSON with: vendor_name, invoice_number, invoice_date, "
        "line_items (list of {product, quantity, unit_price, total}), subtotal, total. "
        "Use null for missing fields."
    )
    return _process(file_data, file_type, prompt)


@mcp.tool()
def extract_purchase_order(file_data: str, file_type: str) -> str:
    prompt = (
        "Extract structured purchase order data. "
        "Return JSON with: vendor_name, po_number, order_date, "
        "line_items (list of {product, quantity, unit_price, total}), subtotal, total. "
        "Use null for missing fields."
    )
    return _process(file_data, file_type, prompt)


@mcp.tool()
def extract_sales_order(file_data: str, file_type: str) -> str:
    prompt = (
        "Extract structured sales order data. "
        "Return JSON with: customer_name, order_number, order_date, "
        "line_items (list of {product, quantity, unit_price, total}), subtotal, total. "
        "Use null for missing fields."
    )
    return _process(file_data, file_type, prompt)


if __name__ == "__main__":
    mcp.run(transport="stdio")
